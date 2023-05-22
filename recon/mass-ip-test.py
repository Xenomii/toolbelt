import subprocess
import openpyxl
import threading

# Define a function to perform the tests for a single IP address
def test_ip(ip, row):
    print(f"Testing {ip}...")

    # Ping test
    ping_result = subprocess.run(["ping", "-c", "3", ip], capture_output=True, text=True)
    if "3 received" in ping_result.stdout:
        ping_outcome = "Success"
    else:
        ping_outcome = "Failure"
    worksheet.cell(row=row, column=2, value=ping_outcome)

    # nslookup
    nslookup_result = subprocess.run(["nslookup", ip], capture_output=True, text=True)
    if "Name:" in nslookup_result.stdout:
        nslookup_outcome = "Success"
    else:
        nslookup_outcome = "Failure"
    worksheet.cell(row=row, column=3, value=nslookup_outcome)

    # Traceroute
    traceroute_result = subprocess.run(["traceroute", "-I", ip], capture_output=True, text=True)
    if ip in traceroute_result.stdout.splitlines()[-1]:
        traceroute_outcome = "Success"
    else:
        traceroute_outcome = "Failure"
    worksheet.cell(row=row, column=4, value=traceroute_outcome)

    # nmap scan
    nmap_result = subprocess.run(["nmap", "-sS", "-p", "21,22,23,25,80,110,139,143,443,445,3389,8080", ip], capture_output=True, text=True)
    if "open" in nmap_result.stdout:
        nmap_outcome = "Success"
    else:
        nmap_outcome = "Failure"
    worksheet.cell(row=row, column=5, value=nmap_outcome)

    # Write the IP address to the first column of the row
    worksheet.cell(row=row, column=1, value=ip)

# Open a new workbook and create a worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Write headers for the columns
worksheet.cell(row=1, column=1, value="IP Address")
worksheet.cell(row=1, column=2, value="Ping Test")
worksheet.cell(row=1, column=3, value="nslookup")
worksheet.cell(row=1, column=4, value="Traceroute")
worksheet.cell(row=1, column=5, value="nmap scan")

# Read the IP address list from a file
with open("ip_list.txt", "r") as f:
    ip_list = [line.strip() for line in f]

    # Create a list to hold the threads
    threads = []

    # Loop through each IP address and start a new thread to perform the tests for that IP
    for i, ip in enumerate(ip_list):
        # Create a new thread for the current IP address
        thread = threading.Thread(target=test_ip, args=(ip, i+2))
        threads.append(thread)

        # Start the thread
        thread.start()

    # Wait for all threads to finish before saving the workbook
    for thread in threads:
        thread.join()

# Save the workbook
workbook.save("test_results.xlsx")

